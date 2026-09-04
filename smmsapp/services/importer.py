"""Bulk student onboarding import logic.

Pure, testable parsing/validation/commit logic for importing a school's students
(and optionally their RFID cards and parent links) from a CSV or XLSX file.

Design:
- parse()   reads a file/bytes and normalizes it into row dicts with row_number.
- validate() returns a row-level report: each row is valid/warning/error with a
  precise list of problems. Never mutates the database.
- commit()  imports the validated rows according to a policy:
    best_effort   -> import valid rows, skip invalid ones
    all_or_nothing-> import all valid rows in one transaction, else roll back
- Only admins call this; each school imports its own students, so the school is
  taken from the importing admin (never from the file).
"""

import csv
import io
import random
from datetime import datetime

from ..models import (
    CustomUser, RFIDCard, ParentStudent, Notification, School,
)

# Expected header names (exact, case-sensitive) and their required-ness.
EXPECTED_COLUMNS = [
    ('first_name', True),
    ('middle_name', False),
    ('last_name', True),
    ('gender', False),
    ('class_room', False),
    ('card_number', False),
    ('parent_email', False),
    ('parent_mobile', False),
]

VALID_GENDERS = {'M', 'F'}


class ImportError(Exception):
    """Raised for structural file errors (missing headers, empty file, etc.)."""


def _generate_control_number(school_number):
    """Recreate the control number used by CreateRFIDCardSerializer."""
    year = datetime.now().year % 100
    month = f"{datetime.now().month:02d}"
    random4 = random.randint(1000, 9999)
    return f"{school_number}{year}{month}{random4}"


def _unique_control_number(school_number):
    """Return a control number that is not already in use."""
    for _ in range(20):
        candidate = _generate_control_number(school_number)
        if not RFIDCard.objects.filter(control_number=candidate).exists():
            return candidate
    # Fallback: include a larger random to avoid a collision.
    return f"{school_number}{datetime.now().year % 100}{datetime.now().month:02d}{random.randint(100000, 999999)}"


class StudentImporter:
    COLUMNS = EXPECTED_COLUMNS

    def __init__(self, school):
        self.school = school

    # ------------------------------------------------------------------ parse
    def parse(self, content, filename=''):
        """Parse raw file bytes into a list of row dicts + header used."""
        ext = (filename or '').lower().rsplit('.', 1)[-1]
        if ext == 'xlsx':
            rows = self._parse_xlsx(content)
        else:
            rows = self._parse_csv(content)
        return rows

    def _parse_csv(self, content):
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        missing = [c for c, _ in self.COLUMNS if c not in (reader.fieldnames or [])]
        if missing:
            raise ImportError(f"Missing required columns: {', '.join(missing)}")
        rows = []
        for i, row in enumerate(reader, start=2):  # row 1 is the header
            rows.append({'row_number': i, 'data': {k: (row.get(k) or '').strip() for k, _ in self.COLUMNS}})
        return rows

    def _parse_xlsx(self, content):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        header = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
        missing = [c for c, _ in self.COLUMNS if c not in header]
        if missing:
            raise ImportError(f"Missing required columns: {', '.join(missing)}")
        col_index = {name: header.index(name) for name, _ in self.COLUMNS}
        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None or str(v).strip() == '' for v in row):
                continue  # skip fully blank rows
            data = {}
            for name, _ in self.COLUMNS:
                val = row[col_index[name]] if col_index[name] < len(row) else None
                data[name] = str(val).strip() if val is not None else ''
            rows.append({'row_number': i, 'data': data})
        return rows

    # ---------------------------------------------------------------- validate
    def validate_rows(self, rows):
        """Return a row-level report (no DB mutation).

        Report entries: {row_number, first_name, status, errors[], parent_ok}
        """
        report = []

        # Caches to detect intra-file duplicates (card_number, username collisions).
        seen_card_numbers = set()
        seen_usernames = set()

        for entry in rows:
            row_number = entry['row_number']
            data = entry['data']
            errors = []

            # --- required fields
            if not data['first_name']:
                errors.append("first_name is required")
            if not data['last_name']:
                errors.append("last_name is required")

            # --- gender
            if data['gender'] and data['gender'].upper() not in VALID_GENDERS:
                errors.append("gender must be 'M' or 'F' (or blank)")

            # --- card_number: uniqueness against DB and intra-file
            card_number = data['card_number']
            if card_number:
                if card_number in seen_card_numbers:
                    errors.append("duplicate card_number within this file")
                elif RFIDCard.objects.filter(card_number=card_number).exists():
                    errors.append("card_number already used on another card")
                else:
                    seen_card_numbers.add(card_number)

            # --- parent resolution (optional; a warning, not an error)
            parent_warnings = []
            parent_mobile = data['parent_mobile']
            parent_email = data['parent_email']
            if parent_mobile or parent_email:
                parent = None
                if parent_email:
                    parent = CustomUser.objects.filter(role='parent', email__iexact=parent_email).first()
                if parent is None and parent_mobile:
                    parent = CustomUser.objects.filter(role='parent', mobile_number=parent_mobile).first()
                if parent is None:
                    parent_warnings.append("no matching parent account found for parent_email/parent_mobile")

            # --- username uniqueness (collision within the file is an error,
            #     since two students would end up with the same generated names).
            if data['first_name'] and data['last_name']:
                key = f"{data['first_name'].lower()}.{data['last_name'].lower()}.{self.school.name.lower()}{datetime.now().year}"
                if key in seen_usernames:
                    errors.append("duplicate student name within this file (username collision)")
                else:
                    seen_usernames.add(key)

            entry_st = 'error' if errors else 'valid'

            report.append({
                'row_number': row_number,
                'first_name': data['first_name'],
                'last_name': data['last_name'],
                'card_number': card_number,
                'status': entry_st,
                'errors': errors,
                'warnings': parent_warnings,
            })
        return report

    # ------------------------------------------------------------------ commit
    def commit_rows(self, rows, report, mode='best_effort'):
        """Import rows according to a policy. Returns a summary.

        mode='best_effort':   import valid rows, skip invalid ones.
        mode='all_or_nothing': all valid rows import in one transaction; if a
                               non-valid row exists the whole batch is rejected.
        """
        report_by_row = {r['row_number']: r for r in report}

        if mode == 'all_or_nothing':
            invalid = [r for r in report if r['status'] != 'valid']
            if invalid:
                return {
                    'committed': [],
                    'skipped': [r['row_number'] for r in invalid],
                    'all_or_nothing_aborted': True,
                    'invalid_count': len(invalid),
                }, report

        committed = []
        for entry in rows:
            rec = report_by_row.get(entry['row_number'], {})
            if rec.get('status') != 'valid':
                continue
            row = entry['data']
            created = self._create_one(row)
            committed.append({
                'row_number': entry['row_number'],
                'student_id': str(created['student'].id),
                'username': created['student'].username,
                'card_id': str(created['card'].id) if created['card'] else None,
                'card_number': created['card'].card_number if created['card'] else None,
                'control_number': created['card'].control_number if created['card'] else None,
            })
        return {'committed': committed, 'skipped': [], 'all_or_nothing_aborted': False}, report

    def _create_one(self, row):
        """Create a student (+ optional card & parent link) for one validated row."""
        from django.db import transaction

        with transaction.atomic():
            # Register the username as taken so later rows won't collide.
            first_name = row['first_name']
            last_name = row['last_name']
            middle_name = row['middle_name']
            gender = row['gender'].upper() if row['gender'] else 'M'
            class_room = row['class_room']
            card_number = row['card_number']

            base = f"{first_name.lower()}.{last_name.lower()}.{self.school.name.lower()}{datetime.now().year}"
            username = base
            if CustomUser.objects.filter(username=username).exists():
                suffix = 1
                while CustomUser.objects.filter(username=f"{base}{suffix}").exists():
                    suffix += 1
                username = f"{base}{suffix}"

            student = CustomUser.objects.create(
                username=username,
                role='student',
                first_name=first_name,
                middle_name=middle_name,
                last_name=last_name,
                gender=gender,
                class_room=class_room or None,
                school=self.school,
            )

            card = None
            if card_number:
                card = RFIDCard.objects.create(
                    card_number=card_number,
                    control_number=_unique_control_number(self.school.number),
                    student_or_staff=student,
                    balance=0.0,
                    is_active=False,  # new cards inactive until activated
                    issued_date=None,
                )
                Notification.objects.create(
                    title=f"{first_name}'s Card Creation",
                    recipient=student,
                    message=f"Your meal card was created. Card Number: {card_number}, "
                            f"Control Number: {card.control_number}, Balance: Tsh. 0.",
                    status='pending',
                    type='reminder',
                )

            # Optional parent link.
            parent_email = row['parent_email']
            parent_mobile = row['parent_mobile']
            if parent_email or parent_mobile:
                parent = None
                if parent_email:
                    parent = CustomUser.objects.filter(role='parent', email__iexact=parent_email).first()
                if parent is None and parent_mobile:
                    parent = CustomUser.objects.filter(role='parent', mobile_number=parent_mobile).first()
                if parent is not None:
                    ParentStudent.objects.get_or_create(parent=parent, student=student)

            return {'student': student, 'card': card}


def build_template_rows():
    """Return a single sample data row for template export."""
    return {
        'first_name': 'John',
        'middle_name': 'Doe',
        'last_name': 'Smith',
        'gender': 'M',
        'class_room': '5A',
        'card_number': 'CARD-000001',
        'parent_email': 'parent@example.com',
        'parent_mobile': '255700000000',
    }


def build_template_csv():
    """Return CSV bytes for the template (header + one sample row)."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=[c for c, _ in EXPECTED_COLUMNS])
    writer.writeheader()
    writer.writerow(build_template_rows())
    return buf.getvalue().encode('utf-8')
