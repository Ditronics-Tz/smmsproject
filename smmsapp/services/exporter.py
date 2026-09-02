import csv
from io import BytesIO, StringIO

from django.conf import settings
from django.db.models import Q

from ..models import Transaction, CustomUser, BankDeposit, CanteenItem
from ..utils import get_admin_scope

EXPORT_SYNC_MAX_ROWS = getattr(settings, 'EXPORT_SYNC_MAX_ROWS', 2000)


# ---------------------------------------------------------------------------
# Role-scoped queryset builders (mirror the list views' filtering
# ---------------------------------------------------------------------------

def transaction_queryset(user, filters):
    qs = Transaction.objects.all().order_by('-transaction_date')

    school = get_admin_scope(user)
    if user.role == 'admin':
        if school is not None:
            qs = qs.filter(student_or_staff__school=school)
    elif user.role == 'parent':
        from ..models import ParentStudent
        children = ParentStudent.objects.filter(parent=user).values_list('student_id', flat=True)
        qs = qs.filter(student_or_staff_id__in=list(children))
    elif user.role == 'staff':
        qs = qs.filter(student_or_staff=user)
    elif user.role == 'operator':
        qs = qs.filter(session__operator=user)
    else:
        qs = qs.none()

    return _apply_transaction_filters(qs, filters)


def _apply_transaction_filters(qs, filters):
    from_date = filters.get('from_date')
    to_date = filters.get('to_date')
    status_val = (filters.get('status') or '').strip()
    search = (filters.get('search') or '').strip()

    if from_date:
        qs = qs.filter(transaction_date__date__gte=from_date)
    if to_date:
        qs = qs.filter(transaction_date__date__lte=to_date)
    if status_val:
        qs = qs.filter(transaction_status=status_val)
    if search:
        qs = qs.filter(
            Q(student_or_staff__username__icontains=search) |
            Q(student_or_staff__first_name__icontains=search) |
            Q(student_or_staff__last_name__icontains=search) |
            Q(rfid_card__card_number__icontains=search)
        )
    return qs


def student_queryset(user, filters):
    qs = CustomUser.objects.filter(role='student').order_by('first_name')

    school = get_admin_scope(user)
    if user.role == 'admin':
        if school is not None:
            qs = qs.filter(school=school)
    elif user.role == 'parent':
        from ..models import ParentStudent
        qs = qs.filter(parents__parent=user)
    else:
        qs = qs.filter(id=user.id)

    search = (filters.get('search') or '').strip()
    class_room = (filters.get('class_room') or '').strip()
    active = filters.get('active')

    if search:
        qs = qs.filter(
            Q(username__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(mobile_number__icontains=search)
        )
    if class_room:
        qs = qs.filter(class_room__icontains=class_room)
    if active is not None:
        qs = qs.filter(is_active=active)
    return qs


def deposit_queryset(user, filters):
    qs = BankDeposit.objects.all().order_by('-created_at')

    school = get_admin_scope(user)
    if user.role in ('admin', 'operator'):
        if school is not None:
            qs = qs.filter(control_number__student_or_staff__school=school)
    elif user.role == 'parent':
        qs = qs.filter(submitted_by=user)
    else:
        qs = qs.none()

    from_date = filters.get('from_date')
    to_date = filters.get('to_date')
    status_val = (filters.get('status') or '').strip()

    if from_date:
        qs = qs.filter(created_at__date__gte=from_date)
    if to_date:
        qs = qs.filter(created_at__date__lte=to_date)
    if status_val:
        qs = qs.filter(status=status_val)
    return qs


# ---------------------------------------------------------------------------
# Row builders -> list of scalar values aligned with headers
# ---------------------------------------------------------------------------

def _transaction_rows(qs):
    rows = []
    for t in qs.select_related('student_or_staff', 'item'):
        rows.append([
            t.id,
            t.student_or_staff.username,
            f"{t.student_or_staff.first_name} {t.student_or_staff.last_name}".strip(),
            t.rfid_card.card_number,
            t.item.name if t.item else '',
            t.amount,
            t.transaction_status,
            t.transaction_date.isoformat() if t.transaction_date else '',
            'Voided' if t.is_voided else '',
        ])
    return rows


def _student_rows(qs):
    rows = []
    for s in qs.select_related('school'):
        card = s.rfid_cards.filter(is_active=True).first()
        rows.append([
            s.username,
            s.first_name,
            s.last_name,
            s.class_room or '',
            s.school.name if s.school else '',
            s.mobile_number or '',
            card.balance if card else '',
            'Active' if s.is_active else 'Inactive',
        ])
    return rows


def _deposit_rows(qs):
    rows = []
    for d in qs.select_related('control_number', 'submitted_by'):
        rows.append([
            d.control_number.card_number,
            d.amount,
            d.status,
            d.created_at.date().isoformat() if d.created_at else '',
            d.processed_at.date().isoformat() if d.processed_at else '',
            d.submitted_by.username if d.submitted_by else '',
        ])
    return rows


TRANSACTION_HEADERS = [
    'Transaction ID', 'Username', 'Name', 'Card Number', 'Item', 'Amount',
    'Status', 'Transaction Date', 'Voided',
]
STUDENT_HEADERS = [
    'Username', 'First Name', 'Last Name', 'Class Room', 'School', 'Mobile',
    'Balance', 'Status',
]
DEPOSIT_HEADERS = [
    'Card Number', 'Amount', 'Status', 'Created Date', 'Processed Date', 'Submitted By',
]

ENTITY_BUILDERS = {
    'transactions': (transaction_queryset, _transaction_rows, TRANSACTION_HEADERS),
    'students': (student_queryset, _student_rows, STUDENT_HEADERS),
    'deposits': (deposit_queryset, _deposit_rows, DEPOSIT_HEADERS),
}


# ---------------------------------------------------------------------------
# Serializers to bytes (CSV / xlsx)
# ---------------------------------------------------------------------------

def export_to_csv(rows, headers):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    buffer.seek(0)
    return buffer.getvalue().encode('utf-8-sig')


def export_to_xlsx(rows, headers):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)

    for idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 22

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
